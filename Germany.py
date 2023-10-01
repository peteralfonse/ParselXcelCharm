#!/usr/bin/env python
# coding: utf-8

# In[1]:


import PyPDF2
import openpyxl
import os
import re
import sys


# In[2]:


OLD_PDF_PATH = r"NEW_PDF_PATH"
OLD_PDF_NAME = os.path.basename(OLD_PDF_PATH)
OLD_PDF_DIR = os.path.dirname(OLD_PDF_PATH)


# In[3]:


with open(OLD_PDF_PATH, 'rb') as pdf_file:
    pdf_reader = PyPDF2.PdfReader(pdf_file)
    
    pdf_text = ''
    for page_num in range(len(pdf_reader.pages)):
        page = pdf_reader.pages[page_num]
        pdf_text += page.extract_text()


# In[4]:


# Split the extracted text into lines
text_lines = pdf_text.split('\n')


# In[5]:


# Create an Excel workbook
excel_file_name = os.path.splitext(os.path.basename(OLD_PDF_PATH))[0] + '.xlsx'
excel_file_path = os.path.join(OLD_PDF_DIR, excel_file_name)

workbook = openpyxl.Workbook()
sheet = workbook.active
sheet.title = "hope"  # Set the title of the first sheet to "hope"


# In[6]:


# Assign each line to a separate row in the "hope" sheet
for row_num, line in enumerate(text_lines, start=1):
    sheet.cell(row=row_num, column=1, value=line)


# In[7]:


sheet = workbook['hope']  # Assuming the sheet name is "hope"

for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=1, max_col=1):
    cell_value = row[0].value
    match = re.search(r'G\d{10}', cell_value)
    if match:
        matched_value = match.group()
        corresponding_cell = sheet.cell(row=row[0].row, column=2)
        corresponding_cell.value = matched_value


# In[8]:


previous_value = None
for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=2, max_col=2):
    cell = row[0]
    if cell.value is None and previous_value is not None:
        cell.value = previous_value
    else:
        previous_value = cell.value


# In[9]:


sheet = workbook['hope']  # Assuming the sheet name is "hope"

for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=1, max_col=1):
    cell_value = row[0].value
    match = re.search(r'\d{10}', cell_value)  # Find 10-digit numbers
    if match and (cell_value.startswith('157') or cell_value.startswith('106')):
        number = match.group()
        corresponding_cell = sheet.cell(row=row[0].row, column=3)
        corresponding_cell.value = number


# In[10]:


sheet = workbook['hope']  # Assuming the sheet name is "hope"

for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=1, max_col=1):
    cell_value = row[0].value
    if cell_value.startswith('PROFORMA'):
        corresponding_cell = sheet.cell(row=row[0].row, column=3)
        corresponding_cell.value = "PROFORMA"


# In[11]:


previous_value = None
for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=3, max_col=3):
    cell = row[0]
    if cell.value is None and previous_value is not None:
        cell.value = previous_value
    else:
        previous_value = cell.value


# In[12]:


sheet = workbook['hope']  # Assuming the sheet name is "hope"

formats = [
    "A 000 000 0000",
    "A 000 000 0000 00",
    "A 000000 000000",
    "A 000000 000000 00",
    "A 000 000 0000 0000",
    "A 000 000 0000 00 0000",
    "A 000000000",
    "A 000000000 00",
    "A 000 000 0000 00 0A00",
    "A 000 000 0000 0A00",
    "A 000000 000000 0000"
]

for i in range(1, sheet.max_row + 1):  # Assuming the header is in the first row
    line = sheet.cell(row=i, column=1).value
    found_format = None
    for format in formats:
        if re.search(format.replace("A", "[A-Za-z]").replace("0", "[0-9]"), line):
            found_format = format
            break
    
    if found_format:
        next_row_line = sheet.cell(row=i + 1, column=1).value
        next_row_found_format = None
        for format in formats:
            if re.search(format.replace("A", "[A-Za-z]").replace("0", "[0-9]"), next_row_line):
                next_row_found_format = format
                break
        
        if not next_row_found_format:
            sheet.cell(row=i, column=4).value = line
            sheet.cell(row=i, column=5).value = sheet.cell(row=i + 1, column=1).value
            sheet.cell(row=i, column=6).value = sheet.cell(row=i + 2, column=1).value
            sheet.cell(row=i, column=7).value = sheet.cell(row=i + 3, column=1).value
            sheet.cell(row=i, column=8).value = sheet.cell(row=i + 4, column=1).value
        else:
            sheet.cell(row=i, column=4).value = line
            sheet.cell(row=i, column=5).value = sheet.cell(row=i + 1, column=1).value
            sheet.cell(row=i, column=6).value = sheet.cell(row=i + 2, column=1).value
            sheet.cell(row=i, column=7).value = sheet.cell(row=i + 3, column=1).value
            sheet.cell(row=i, column=8).value = sheet.cell(row=i + 4, column=1).value
            sheet.cell(row=i, column=9).value = sheet.cell(row=i + 5, column=1).value


# In[13]:


# Assuming the Excel workbook is already loaded and named 'workbook'
sheet = workbook['hope']  # Assuming the sheet name is "hope"

for i in range(1, sheet.max_row + 1):  # Assuming the header is in the first row
    cell_value = sheet.cell(row=i, column=1).value
    last_space_index = cell_value.rfind(" ")
    if last_space_index != -1:
        value_after_last_space = cell_value[last_space_index + 1:]
        sheet.cell(row=i, column=10).value = value_after_last_space


# In[14]:


sheet = workbook['hope']  # Assuming the sheet name is "hope"

for i in range(1, sheet.max_row + 1):  # Assuming the header is in the first row
    cell_value = sheet.cell(row=i, column=1).value
    last_space_index = cell_value.rfind(" ")
    second_last_space_index = cell_value.rfind(" ", 0, last_space_index)
    
    if last_space_index != -1 and second_last_space_index != -1:
        value_between_spaces = cell_value[second_last_space_index + 1:last_space_index]
        sheet.cell(row=i, column=11).value = value_between_spaces


# In[15]:


sheet = workbook['hope']  # Assuming the sheet name is "hope"

for i in range(1, sheet.max_row + 1):  # Assuming the header is in the first row
    cell_value = sheet.cell(row=i, column=1).value
    last_space_index = cell_value.rfind(" ")
    second_last_space_index = cell_value.rfind(" ", 0, last_space_index)
    third_last_space_index = cell_value.rfind(" ", 0, second_last_space_index)
    
    if second_last_space_index != -1 and third_last_space_index != -1:
        value_between_spaces = cell_value[third_last_space_index + 1:second_last_space_index]
        sheet.cell(row=i, column=12).value = value_between_spaces


# In[16]:


sheet = workbook['hope']  # Assuming the sheet name is "hope"

for i in range(1, sheet.max_row + 1):  # Assuming the header is in the first row
    cell_value = sheet.cell(row=i, column=1).value
    last_space_index = cell_value.rfind(" ")
    second_last_space_index = cell_value.rfind(" ", 0, last_space_index)
    third_last_space_index = cell_value.rfind(" ", 0, second_last_space_index)
    fourth_last_space_index = cell_value.rfind(" ", 0, third_last_space_index)
    
    if third_last_space_index != -1 and fourth_last_space_index != -1:
        value_between_spaces = cell_value[fourth_last_space_index + 1:third_last_space_index]
        sheet.cell(row=i, column=13).value = value_between_spaces


# In[17]:


sheet = workbook['hope']  # Assuming the sheet name is "hope"

for i in range(1, sheet.max_row + 1):  # Assuming the header is in the first row
    cell_value = sheet.cell(row=i, column=1).value
    last_space_index = cell_value.rfind(" ")
    second_last_space_index = cell_value.rfind(" ", 0, last_space_index)
    third_last_space_index = cell_value.rfind(" ", 0, second_last_space_index)
    fourth_last_space_index = cell_value.rfind(" ", 0, third_last_space_index)
    fifth_last_space_index = cell_value.rfind(" ", 0, fourth_last_space_index)
    
    if fourth_last_space_index != -1 and fifth_last_space_index != -1:
        value_between_spaces = cell_value[fifth_last_space_index + 1:fourth_last_space_index]
        sheet.cell(row=i, column=14).value = value_between_spaces


# In[18]:


sheet = workbook['hope']  # Assuming the sheet name is "hope"

for i in range(1, sheet.max_row + 1):  # Assuming the header is in the first row
    cell_value = sheet.cell(row=i, column=1).value
    first_space_index = cell_value.find(" ")
    sixth_last_space_index = cell_value.rfind(" ", 0, cell_value.rfind(" ", 0, cell_value.rfind(" ", 0, cell_value.rfind(" ", 0, cell_value.rfind(" ") - 1) - 1) - 1) - 1)
    
    if first_space_index != -1 and sixth_last_space_index != -1:
        value_between_spaces = cell_value[first_space_index + 1:sixth_last_space_index]
        sheet.cell(row=i, column=14).value = value_between_spaces  # Change to column 14 (column N)


# In[19]:


sheet = workbook['hope']  # Assuming the sheet name is "hope"

# Iterate through rows in column "A"
for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=1, max_col=1):
    cell_value = row[0].value
    space_indices = [pos for pos, char in enumerate(cell_value) if char == ' ']
    
    if len(space_indices) >= 5:
        fifth_last_space_index = space_indices[-5]
        fourth_last_space_index = space_indices[-4]
        value_between_spaces = cell_value[fifth_last_space_index + 1:fourth_last_space_index]
        sheet.cell(row=row[0].row, column=15).value = value_between_spaces


# In[20]:


#AFTER THIS STEP WE WILL START DELETING AND CLEANING


# In[21]:


sheet = workbook['hope']  # Assuming the sheet name is "hope"

rows_to_delete = []  # Store row indices to delete

for i in range(sheet.max_row, 0, -1):  # Loop in reverse to avoid index changes
    if sheet.cell(row=i, column=4).value is None:
        rows_to_delete.append(i)

for row_index in rows_to_delete:
    sheet.delete_rows(row_index)


# In[22]:


sheet = workbook['hope']  # Assuming the sheet name is "hope"

# Insert a new column between columns "D" and "E"
sheet.insert_cols(5)


# In[23]:


sheet = workbook['hope']  # Assuming the sheet name is "hope"

rows_to_shift = []  # Store row indices to shift

for i in range(1, sheet.max_row + 1):
    if sheet.cell(row=i, column=10).value is not None:  # Checking column "J"
        rows_to_shift.append(i)

for row_index in rows_to_shift:
    # Shift cells in columns "F" to "J" to columns "E" to "I"
    for col in range(6, 11):
        source_cell_value = sheet.cell(row=row_index, column=col).value
        sheet.cell(row=row_index, column=col - 1).value = source_cell_value
        sheet.cell(row=row_index, column=col).value = None


# In[24]:


sheet = workbook['hope']  # Assuming the sheet name is "hope"

columns_to_delete = [1, 4, 10]  # Column indices to delete

for col_index in sorted(columns_to_delete, reverse=True):
    sheet.delete_cols(col_index)


# In[25]:


sheet = workbook['hope']  # Assuming the sheet name is "hope"

rows_to_delete = []

for i in range(1, sheet.max_row + 1):
    cell_value = sheet.cell(row=i, column=12).value  # Column L
    if cell_value is None or cell_value.strip() == '':
        rows_to_delete.append(i)

for row_index in reversed(rows_to_delete):  # Deleting in reverse order to avoid index changes
    sheet.delete_rows(row_index)


# In[26]:


sheet = workbook['hope']  # Assuming the sheet name is "hope"

for i in range(1, sheet.max_row + 1):  # Assuming the header is in the first row
    cell_value = sheet.cell(row=i, column=3).value  # Column C
    
    if cell_value is not None:
        first_space_index = cell_value.find(" ")
        last_space_index = cell_value.rfind(" ")
    
        if first_space_index != -1 and last_space_index != -1:
            value_between_spaces = cell_value[first_space_index + 1:last_space_index]
            sheet.cell(row=i, column=14).value = value_between_spaces  # Column N


# In[27]:


sheet = workbook['hope']  # Assuming the sheet name is "hope"

column_to_delete = 14  # Column N index

sheet.delete_cols(column_to_delete)


# In[28]:


sheet = workbook['hope']  # Assuming the sheet name is "hope"

# Column names for the new header
new_header = [
    "Delivery Note",
    "Invoice Number",
    "Alternative Part Number",
    "Description Deutsch",
    "Description English",
    "CRM",
    "Order Number",
    "Discount Group",
    "Invoice Value",
    "Total Price",
    "Unit Price",
    "Part Number",
    "Quantity"
]

# Insert the new header above the first row
sheet.insert_rows(1)

# Populate the new header with column names
for col_index, col_name in enumerate(new_header, start=1):
    sheet.cell(row=1, column=col_index).value = col_name


# In[29]:


# Assuming you've already loaded the workbook and selected the "hope" sheet
sheet = workbook['hope']

# Step 1: Remove "." from columns I, J, K
for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=9, max_col=11):
    for cell in row:
        if cell.value is not None and isinstance(cell.value, str):
            cell.value = cell.value.replace(".", "")

# Step 2: Replace "," with "." in columns I, J, K
for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=9, max_col=11):
    for cell in row:
        if cell.value is not None and isinstance(cell.value, str):
            cell.value = cell.value.replace(",", ".")


# In[30]:


sheet = workbook['hope']  # Assuming the sheet name is "hope"

# Rearrange the columns
column_order = [
    "L", "C", "M", "I", "J", "H", "E", "A", "B", "G", "F", "K", "D"
]

# Create a new worksheet to store reordered data
reordered_sheet = workbook.create_sheet(title='reordered_hope', index=0)

# Copy data to the reordered sheet in the specified order
for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row):
    reordered_row = [row[sheet[col_name + "1"].column - 1].value for col_name in column_order]
    reordered_sheet.append(reordered_row)

# Delete the original "hope" sheet
workbook.remove(sheet)

# Rename the reordered sheet to "hope"
reordered_sheet.title = "hope"


# In[31]:


sheet = workbook.active  # Assuming you want to work with the active sheet

# Get the new sheet name from the first value in column J
new_sheet_name = sheet.cell(row=2, column=10).value  # Assuming row 1 is the header

# Rename the sheet
sheet.title = new_sheet_name


# In[32]:


sheet = workbook.active  # Assuming you want to work with the active sheet

formats = [
    "A 000 000 0000",
    "A 000 000 0000 00",
    "A 000000 000000",
    "A 000000 000000 00",
    "A 000 000 0000 0000",
    "A 000 000 0000 00 0000",
    "A 000000000",
    "A 000000000 00",
    "A 000 000 0000 00 0A00",
    "A 000 000 0000 0A00",
    "A 000000 000000 0000"
]

rows_to_delete = []

# Iterate through rows in column A and check for valid formats
for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=1):
    cell_value = row[0].value
    found_valid_format = False
    for format_pattern in formats:
        if re.match(format_pattern.replace("A", "[A-Za-z]").replace("0", "[0-9]"), cell_value):
            found_valid_format = True
            break
    
    if not found_valid_format:
        rows_to_delete.append(row[0].row)

# Delete the rows with invalid formats
for row_index in reversed(rows_to_delete):
    sheet.delete_rows(row_index)


# In[33]:


try:
    # Save the Excel file in the same path as the PDF
    workbook.save(excel_file_path)
    print(f"Excel file '{excel_file_name}' created with PDF text.")
except Exception as e:
    print(f"An error occurred while saving the Excel file: {e}")


print(f"PDF file path: {OLD_PDF_PATH}")
print(f"Excel file path: {excel_file_path}")


# In[ ]:




