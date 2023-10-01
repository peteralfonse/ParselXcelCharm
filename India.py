#!/usr/bin/env python
# coding: utf-8

# In[38]:


import os
import re
import PyPDF2
import openpyxl
from openpyxl import Workbook  # Import Workbook here
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font
import sys


# In[39]:

OLD_PDF_PATH = r"NEW_PDF_PATH"
OLD_PDF_NAME = os.path.basename(OLD_PDF_PATH)
OLD_PDF_DIR = os.path.dirname(OLD_PDF_PATH)


# In[40]:


lines = []

with open(OLD_PDF_PATH, 'rb') as pdf_file:
    pdf_reader = PyPDF2.PdfReader(pdf_file)
    
    for page in pdf_reader.pages:
        text = page.extract_text()
        lines.extend(text.split('\n'))

exe_dir = getattr(sys, '_MEIPASS', os.path.abspath(os.path.dirname(__file__)))
# Create an Excel workbook
excel_file_name = os.path.splitext(os.path.basename(OLD_PDF_PATH))[0] + '.xlsx'
excel_file_path = os.path.join(OLD_PDF_DIR, excel_file_name)
# Create an Excel workbook
# excel_file_name = os.path.splitext(os.path.basename(OLD_PDF_PATH))[0] + '.xlsx'
# excel_file_path = os.path.join(os.path.dirname(OLD_PDF_PATH), excel_file_name)
wb = Workbook()
ws = wb.active

for line in lines:
    ws.append([line])

wb.save(excel_file_path)
print("Excel sheet created successfully.")


# In[41]:


format_regexes = [r' [A-Za-z]{1}[0-9]{10} ', r' [A-Za-z]{2}[0-9]{6} ']

for row in ws.iter_rows():
    for cell in row:
        cell_text = cell.value
        if cell_text:
            for format_regex in format_regexes:
                matches = re.findall(format_regex, cell_text)
                if matches:
                    fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
                    cell.fill = fill

wb.save(excel_file_path)


# In[42]:


# Remove trailing spaces in column A
for row in ws.iter_rows():
    if row[0].value:
        row[0].value = row[0].value.rstrip()
        
wb.save(excel_file_path)


# In[43]:


format_regexes = [r' [A-Za-z]{1}[0-9]{10} ', r' [A-Za-z]{2}[0-9]{6} ']

# Iterate through rows in column A
for row_index in range(1, ws.max_row + 1):
    cell_A = ws.cell(row=row_index, column=1)
    if cell_A.value:
        cell_A.value = cell_A.value.rstrip()
        if any(re.search(format_regex, cell_A.value) for format_regex in format_regexes) and not re.search(r'\d$', cell_A.value):
            row_below = ws.cell(row=row_index + 1, column=1)
            if row_below.value:
                cell_B = ws.cell(row=row_index, column=2)
                cell_B.value = row_below.value

wb.save(excel_file_path)


# In[44]:


format_regexes = [r' [A-Za-z]{1}[0-9]{10} ', r' [A-Za-z]{2}[0-9]{6} ']

# Iterate through rows in column A
for row_index in range(1, ws.max_row + 1):
    cell_A = ws.cell(row=row_index, column=1)
    if cell_A.value:
        cell_A.value = cell_A.value.rstrip()
        if any(re.search(format_regex, cell_A.value) for format_regex in format_regexes) and not re.search(r'\d$', cell_A.value):
            row_below = ws.cell(row=row_index + 1, column=1)
            if row_below.value:
                cell_B = ws.cell(row=row_index, column=2)
                cell_B.value = row_below.value

# Check and copy cells from column A to column C
for row_index in range(1, ws.max_row + 1):
    cell_B = ws.cell(row=row_index, column=2)
    if cell_B.value and not re.search(r'\d$', cell_B.value):
        cell_A_below = ws.cell(row=row_index + 2, column=1)
        cell_C = ws.cell(row=row_index, column=3)
        cell_C.value = cell_A_below.value

wb.save(excel_file_path)


# In[45]:


format_regexes = [r' [A-Za-z]{1}[0-9]{10} ', r' [A-Za-z]{2}[0-9]{6} ']

# Iterate through rows in column A
for row_index in range(1, ws.max_row + 1):
    cell_A = ws.cell(row=row_index, column=1)
    if cell_A.value:
        cell_A.value = cell_A.value.rstrip()
        if any(re.search(format_regex, cell_A.value) for format_regex in format_regexes) and not re.search(r'\d$', cell_A.value):
            row_below = ws.cell(row=row_index + 1, column=1)
            if row_below.value:
                cell_B = ws.cell(row=row_index, column=2)
                cell_B.value = row_below.value

# Check and copy cells from column A to column C
for row_index in range(1, ws.max_row + 1):
    cell_B = ws.cell(row=row_index, column=2)
    if cell_B.value and not re.search(r'\d$', cell_B.value):
        cell_A_below = ws.cell(row=row_index + 2, column=1)
        cell_C = ws.cell(row=row_index, column=3)
        cell_C.value = cell_A_below.value

# Check and copy cells from column C to column D
for row_index in range(1, ws.max_row + 1):
    cell_C = ws.cell(row=row_index, column=3)
    if cell_C.value and not re.search(r'\d$', cell_C.value):
        cell_A_below = ws.cell(row=row_index + 2, column=1)
        cell_D = ws.cell(row=row_index, column=4)
        cell_D.value = cell_A_below.value

wb.save(excel_file_path)


# In[46]:


format_regexes = [r' [A-Za-z]{1}[0-9]{10} ', r' [A-Za-z]{2}[0-9]{6} ']

# Iterate through rows in column A
for row_index in range(1, ws.max_row + 1):
    cell_A = ws.cell(row=row_index, column=1)
    if cell_A.value:
        cell_A.value = cell_A.value.rstrip()
        if any(re.search(format_regex, cell_A.value) for format_regex in format_regexes) and not re.search(r'\d$', cell_A.value):
            row_below = ws.cell(row=row_index + 1, column=1)
            if row_below.value:
                cell_B = ws.cell(row=row_index, column=2)
                cell_B.value = row_below.value

# Check and copy cells from column A to column C
for row_index in range(1, ws.max_row + 1):
    cell_B = ws.cell(row=row_index, column=2)
    if cell_B.value and not re.search(r'\d$', cell_B.value):
        cell_A_below = ws.cell(row=row_index + 2, column=1)
        cell_C = ws.cell(row=row_index, column=3)
        cell_C.value = cell_A_below.value

# Check and copy cells from column C to column D
for row_index in range(1, ws.max_row + 1):
    cell_C = ws.cell(row=row_index, column=3)
    if cell_C.value and not re.search(r'\d$', cell_C.value):
        cell_A_below = ws.cell(row=row_index + 2, column=1)
        cell_D = ws.cell(row=row_index, column=4)
        cell_D.value = cell_A_below.value

# Check and copy cells from columns A, B, C, and D to column E
for row_index in range(1, ws.max_row + 1):
    cell_A = ws.cell(row=row_index, column=1)
    cell_B = ws.cell(row=row_index, column=2)
    cell_C = ws.cell(row=row_index, column=3)
    cell_D = ws.cell(row=row_index, column=4)
    if any(re.search(format_regex, cell_A.value) for format_regex in format_regexes):
        concatenated_value = ''
        if cell_A.value:
            concatenated_value += cell_A.value
        if cell_B.value:
            concatenated_value += cell_B.value
        if cell_C.value:
            concatenated_value += cell_C.value
        if cell_D.value:
            concatenated_value += cell_D.value
        ws.cell(row=row_index, column=5).value = concatenated_value

wb.save(excel_file_path)


# In[47]:


# Delete columns B, C, and D
ws.delete_cols(2, 3)

wb.save(excel_file_path)


# In[48]:


# Iterate through rows and clean values in columns A and B
for row_index in range(1, ws.max_row + 1):
    cell_A = ws.cell(row=row_index, column=1)
    if cell_A.value:
        cell_A.value = re.sub(r'\s+', ' ', cell_A.value).strip()

    cell_B = ws.cell(row=row_index, column=2)
    if cell_B.value:
        cell_B.value = re.sub(r'\s+', ' ', cell_B.value).strip()

    # Extract value from second last space to last space in column B and add to column J
    if cell_B.value:
        space_indices = [i for i, char in enumerate(cell_B.value) if char == ' ']
        if len(space_indices) >= 2:
            second_last_space_index = space_indices[-2]
            last_space_index = space_indices[-1]
            value_from_B = cell_B.value[second_last_space_index+1:last_space_index]
            ws.cell(row=row_index, column=10).value = value_from_B

wb.save(excel_file_path)


# In[49]:


# Iterate through rows and copy value from column B to column K
for row_index in range(1, ws.max_row + 1):
    cell_B = ws.cell(row=row_index, column=2)
    if cell_B.value:
        space_index = cell_B.value.rfind(' ')
        if space_index != -1:
            value_from_B = cell_B.value[space_index+1:]
            ws.cell(row=row_index, column=11).value = value_from_B

wb.save(excel_file_path)


# In[50]:


# Iterate through rows and copy value from second last space to last space in column B to column J
for row_index in range(1, ws.max_row + 1):
    cell_B = ws.cell(row=row_index, column=2)
    if cell_B.value:
        last_space_index = cell_B.value.rfind(' ')
        second_last_space_index = cell_B.value.rfind(' ', 0, last_space_index)
        if second_last_space_index != -1:
            value_from_B = cell_B.value[second_last_space_index+1:last_space_index]
            ws.cell(row=row_index, column=10).value = value_from_B

wb.save(excel_file_path)


# In[51]:


# Iterate through rows and copy value from third last space to second last space in column B to column I
for row_index in range(1, ws.max_row + 1):
    cell_B = ws.cell(row=row_index, column=2)
    if cell_B.value:
        space_indices = [i for i, char in enumerate(cell_B.value) if char == ' ']
        if len(space_indices) >= 3:
            third_last_space_index = space_indices[-3]
            second_last_space_index = space_indices[-2]
            value_from_B = cell_B.value[third_last_space_index+1:second_last_space_index]
            ws.cell(row=row_index, column=9).value = value_from_B

wb.save(excel_file_path)


# In[52]:


# Iterate through rows and copy value from fourth last space to third last space in column B to column H
for row_index in range(1, ws.max_row + 1):
    cell_B = ws.cell(row=row_index, column=2)
    if cell_B.value:
        space_indices = [i for i, char in enumerate(cell_B.value) if char == ' ']
        if len(space_indices) >= 4:
            fourth_last_space_index = space_indices[-4]
            third_last_space_index = space_indices[-3]
            value_from_B = cell_B.value[fourth_last_space_index+1:third_last_space_index]
            ws.cell(row=row_index, column=8).value = value_from_B

wb.save(excel_file_path)


# In[53]:


# Iterate through rows and copy value from fifth last space to fourth last space in column B to column G
for row_index in range(1, ws.max_row + 1):
    cell_B = ws.cell(row=row_index, column=2)
    if cell_B.value:
        space_indices = [i for i, char in enumerate(cell_B.value) if char == ' ']
        if len(space_indices) >= 5:
            fifth_last_space_index = space_indices[-5]
            fourth_last_space_index = space_indices[-4]
            value_from_B = cell_B.value[fifth_last_space_index+1:fourth_last_space_index]
            ws.cell(row=row_index, column=7).value = value_from_B

wb.save(excel_file_path)


# In[54]:


# Iterate through rows and copy value from sixth last space to fifth last space in column B to column F
for row_index in range(1, ws.max_row + 1):
    cell_B = ws.cell(row=row_index, column=2)
    if cell_B.value:
        space_indices = [i for i, char in enumerate(cell_B.value) if char == ' ']
        if len(space_indices) >= 6:
            sixth_last_space_index = space_indices[-6]
            fifth_last_space_index = space_indices[-5]
            value_from_B = cell_B.value[sixth_last_space_index+1:fifth_last_space_index]
            ws.cell(row=row_index, column=6).value = value_from_B

wb.save(excel_file_path)


# In[55]:


# Iterate through rows and copy value from ninth last space to sixth last space in column B to column E
for row_index in range(1, ws.max_row + 1):
    cell_B = ws.cell(row=row_index, column=2)
    if cell_B.value:
        space_indices = [i for i, char in enumerate(cell_B.value) if char == ' ']
        if len(space_indices) >= 9:
            ninth_last_space_index = space_indices[-9]
            sixth_last_space_index = space_indices[-6]
            value_from_B = cell_B.value[ninth_last_space_index+1:sixth_last_space_index]
            ws.cell(row=row_index, column=5).value = value_from_B

wb.save(excel_file_path)


# In[56]:


# Iterate through rows and copy value between second space and ninth last space in column B to column D
for row_index in range(1, ws.max_row + 1):
    cell_B = ws.cell(row=row_index, column=2)
    if cell_B.value:
        space_indices = [i for i, char in enumerate(cell_B.value) if char == ' ']
        if len(space_indices) >= 9:
            second_space_index = space_indices[1]
            ninth_last_space_index = space_indices[-9]
            value_from_B = cell_B.value[second_space_index+1:ninth_last_space_index]
            ws.cell(row=row_index, column=4).value = value_from_B

wb.save(excel_file_path)


# In[57]:


# Iterate through rows and copy value from first space to second space in column B to column C
for row_index in range(1, ws.max_row + 1):
    cell_B = ws.cell(row=row_index, column=2)
    if cell_B.value:
        space_indices = [i for i, char in enumerate(cell_B.value) if char == ' ']
        if len(space_indices) >= 2:
            first_space_index = space_indices[0]
            second_space_index = space_indices[1]
            value_from_B = cell_B.value[first_space_index+1:second_space_index]
            ws.cell(row=row_index, column=3).value = value_from_B

wb.save(excel_file_path)


# In[58]:


# Insert a new column between E and F
ws.insert_cols(6)

wb.save(excel_file_path)


# In[59]:


# Iterate through rows and perform operations based on column G values
for row_index in range(1, ws.max_row + 1):
    cell_G = ws.cell(row=row_index, column=7)
    cell_E = ws.cell(row=row_index, column=5)
    cell_F = ws.cell(row=row_index, column=6)

    if cell_G.value and len(cell_G.value) == 2:
        last_7_chars = cell_E.value[-7:]
        ws.cell(row=row_index, column=5).value = cell_E.value[:-7]

        if cell_F.value is None:
            cell_F.value = ''

        ws.cell(row=row_index, column=6).value = last_7_chars + cell_F.value
        ws.cell(row=row_index, column=5).value += cell_G.value

wb.save(excel_file_path)


# In[60]:


# Iterate through rows and perform operations based on column F values
for row_index in range(1, ws.max_row + 1):
    cell_F = ws.cell(row=row_index, column=6)
    cell_E = ws.cell(row=row_index, column=5)

    if cell_F.value:
        last_2_chars_E = cell_E.value[-2:]
        cell_F.value += last_2_chars_E
        cell_E.value = cell_E.value[:-2]

wb.save(excel_file_path)


# In[61]:


# Iterate through rows and perform operations based on column F values
for row_index in range(1, ws.max_row + 1):
    cell_F = ws.cell(row=row_index, column=6)
    cell_E = ws.cell(row=row_index, column=5)
    cell_D = ws.cell(row=row_index, column=4)

    if cell_F.value:
        cell_D.value += cell_E.value
        cell_E.value = cell_F.value
        cell_F.value = None  # Clear the cell in column F

wb.save(excel_file_path)


# In[62]:


# Delete column F
ws.delete_cols(6)

wb.save(excel_file_path)


# In[63]:


# Iterate through rows and extract values after "Order No.:" in column A
for row_index in range(1, ws.max_row + 1):
    cell_A = ws.cell(row=row_index, column=1)
    cell_O = ws.cell(row=row_index, column=15)  # Column O

    if cell_A.value and cell_A.value.startswith("Invoi ce No :"):
        order_no_text = "Order No.:"
        order_no_index = cell_A.value.find(order_no_text)
        
        if order_no_index != -1:
            order_no_value = cell_A.value[order_no_index + len(order_no_text):].strip()
            cell_O.value = order_no_value

wb.save(excel_file_path)


# In[64]:


# Iterate through rows and extract values between "Date:" and "Order No.:" in column A
for row_index in range(1, ws.max_row + 1):
    cell_A = ws.cell(row=row_index, column=1)
    cell_N = ws.cell(row=row_index, column=14)  # Column N

    if cell_A.value and cell_A.value.startswith("Invoi ce No :"):
        date_text = "Date:"
        order_no_text = "Order No.:"
        date_index = cell_A.value.find(date_text)
        order_no_index = cell_A.value.find(order_no_text)
        
        if date_index != -1 and order_no_index != -1:
            date_value = cell_A.value[date_index + len(date_text):order_no_index].strip()
            cell_N.value = date_value

wb.save(excel_file_path)


# In[65]:


## Iterate through rows and extract values between "Invoi ce No :" and "Date:" in column A
for row_index in range(1, ws.max_row + 1):
    cell_A = ws.cell(row=row_index, column=1)
    cell_M = ws.cell(row=row_index, column=13)  # Column M

    if cell_A.value and cell_A.value.startswith("Invoi ce No :"):
        invoice_text = "Invoi ce No :"
        date_text = "Date:"
        invoice_index = cell_A.value.find(invoice_text)
        date_index = cell_A.value.find(date_text)
        
        if invoice_index != -1 and date_index != -1:
            invoice_number = cell_A.value[invoice_index + len(invoice_text):date_index].strip()
            cell_M.value = invoice_number

wb.save(excel_file_path)


# In[66]:


# Columns to process: M, N, O
columns_to_process = [13, 14, 15]

for col in columns_to_process:
    last_value = None  # Initialize the last value

    # Iterate through rows and fill empty cells with the last non-empty value
    for row_index in range(1, ws.max_row + 1):
        current_cell = ws.cell(row=row_index, column=col)

        if current_cell.value is not None:
            last_value = current_cell.value  # Update the last value
        elif last_value is not None:
            current_cell.value = last_value

wb.save(excel_file_path)


# In[67]:


# Delete columns L and F
columns_to_delete = [6]  # Column L is 12, Column F is 6

for col in reversed(columns_to_delete):
    ws.delete_cols(col)

wb.save(excel_file_path)


# In[68]:


# Delete columns L and F
columns_to_delete = [11]  # Column L is 12, Column F is 6

for col in reversed(columns_to_delete):
    ws.delete_cols(col)

wb.save(excel_file_path)


# In[69]:


# Iterate through rows in reverse order and delete rows without a format in column A
for row_index in range(ws.max_row, 1, -1):
    cell_A = ws.cell(row=row_index, column=1)
    has_format = any(re.search(format_regex, cell_A.value) for format_regex in format_regexes)

    if not has_format:
        ws.delete_rows(row_index)

wb.save(excel_file_path)


# In[70]:


# Iterate through rows and perform operations based on column E values
for row_index in range(1, ws.max_row + 1):
    cell_E = ws.cell(row=row_index, column=5)
    cell_D = ws.cell(row=row_index, column=4)
    
    if cell_E.value and cell_E.value[-1].isalpha():
        extracted_numbers = re.sub(r'\D', '', cell_D.value)
        first_two_chars = cell_E.value[:2]
        
        ws.cell(row=row_index, column=5).value = first_two_chars + extracted_numbers
        ws.cell(row=row_index, column=4).value = cell_D.value.replace(extracted_numbers, '', 1).strip()

wb.save(excel_file_path)


# In[71]:


# Delete column A
ws.delete_cols(1)
# Delete column B
ws.delete_cols(1)

wb.save(excel_file_path)


# In[72]:


# Define the header column names
header_names = [
    "Part Number", "Description", "HSN Code", "Quantity", "Unit Price",
    "Freight", "Insurance", "Total Price", "Invoice Number", "Date", "Order Number"
]

# Insert a new row at the beginning for the header
ws.insert_rows(1)
header_row = ws[1]

# Set the header values and apply styling
for col_idx, header_name in enumerate(header_names, start=1):
    cell = header_row[col_idx - 1]
    cell.value = header_name
    cell.font = Font(bold=True)

# Save the modified workbook
wb.save(excel_file_path)


# In[73]:


# Iterate through rows in reverse order and delete rows with empty or spaces-only cells in column A
for row in reversed(list(ws.iter_rows(min_row=2))):
    if not row[0].value or row[0].value.isspace():
        ws.delete_rows(row[0].row)

# Save the modified workbook
wb.save(excel_file_path)


# In[74]:


ws = wb.active

# Rearrange Columns
new_order_indices = [0, 1, 3, 10, 8, 4, 5, 6, 7, 2, 9]

# Get the column headers before rearranging
original_headers = [cell.value for cell in ws[1]]

for row in ws.iter_rows(min_row=2):
    new_row = [row[idx].value for idx in new_order_indices]
    for col_idx, value in enumerate(new_row, start=1):
        ws.cell(row=row[0].row, column=col_idx, value=value)

# Rearrange Column Headers based on new order
new_headers = [original_headers[idx] for idx in new_order_indices]

# Update header row with new headers
for col_idx, header_name in enumerate(new_headers, start=1):
    ws.cell(row=1, column=col_idx, value=header_name)

# Save the Modified Workbook
wb.save(excel_file_path)


# In[ ]:


try:
    # Save the Excel file in the same path as the PDF
    workbook.save(excel_file_path)
    print(f"Excel file '{excel_file_name}' created with PDF text.")
except Exception as e:
    print(f"An error occurred while saving the Excel file: {e}")


# In[32]:


print(f"PDF file path: {OLD_PDF_PATH}")
print(f"Excel file path: {excel_file_path}")


# In[ ]:




