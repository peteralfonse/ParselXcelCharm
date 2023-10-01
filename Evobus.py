#!/usr/bin/env python
# coding: utf-8

# In[26]:


import os
import PyPDF2
import openpyxl
from openpyxl import Workbook
import re
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows
import sys


# In[27]:

OLD_PDF_PATH = r"NEW_PDF_PATH"
OLD_PDF_NAME = os.path.basename(OLD_PDF_PATH)
OLD_PDF_DIR = os.path.dirname(OLD_PDF_PATH)



pdf_data = ""
with open(OLD_PDF_PATH, "rb") as pdf_file:
    pdf_reader = PyPDF2.PdfReader(pdf_file)
    for page_num in range(len(pdf_reader.pages)):
        pdf_data += pdf_reader.pages[page_num].extract_text()

exe_dir = getattr(sys, '_MEIPASS', os.path.abspath(os.path.dirname(__file__)))
# Create an Excel workbook
excel_file_name = os.path.splitext(os.path.basename(OLD_PDF_PATH))[0] + '.xlsx'
excel_file_path = os.path.join(OLD_PDF_DIR, excel_file_name)
# Create an Excel workbook
# excel_file_name = os.path.splitext(os.path.basename(OLD_PDF_PATH))[0] + '.xlsx'
# excel_file_path = os.path.join(os.path.dirname(OLD_PDF_PATH), excel_file_name)

wb = Workbook()
ws = wb.active

lines = pdf_data.split("\n")
for row_num, line in enumerate(lines, start=1):
    ws.cell(row=row_num, column=1, value=line)




wb.save(excel_file_path)
print(f"Converted {OLD_PDF_PATH} to Excel.")


# In[28]:


formats = [
    (r"A\.\d{3}\.\d{3}\.\d{2}\.\d{2}", "A.000.000.00.00"),
    (r"A\.\d{3}\.\d{3}\.\d{2}\.\d{2}\.\d{2}", "A.000.000.00.00.00"),
    (r"A\.\d{6}\.\d{6}", "A.000000.000000"),
    (r"A\.\d{6}\.\d{6}\.\d{2}", "A.000000.000000.00"),
    (r"A\.\d{3}\.\d{3}\.\d{2}\.\d{2}\.\d{4}", "A.000.000.00.00.0000"),
    (r"A\.\d{3}\.\d{3}\.\d{2}\.\d{2}\.\d{2}\.\d{4}", "A.000.000.00.00.00.0000"),
    (r"A\d{9}", "A.000000000"),
    (r"A\d{9}\.\d{2}", "A.000000000.00"),
    (r"A\.\d{6}\.\d{6}\.\d{4}", "A.000000.000000.0000"),
    (r"A\.\d{3}\.\d{3}\.\d{2}\.\d{2}\.\d[A-Z]00", "A.000.000.00.00.0A00"),
    (r"A\.\d{3}\.\d{3}\.\d{2}\.\d{2}\.\d{2}\.\d[A-Z]00", "A.000.000.00.00.00.0A00")
]


# In[29]:


# Remove rows in column A that contain "_"
rows_to_delete = []
for row_num in range(1, len(lines) + 1):
    if "_" in ws.cell(row=row_num, column=1).value:
        rows_to_delete.append(row_num)

for row_num in reversed(rows_to_delete):
    ws.delete_rows(row_num)
    
wb.save(excel_file_path)


# In[30]:


# Copy cells from column A to column B if they start with the specified text
for row_num in range(2, len(lines) + 1):
    cell_value = ws.cell(row=row_num, column=1).value
    if cell_value and cell_value.startswith("From order/delivery note"):
        ws.cell(row=row_num, column=2).value = cell_value
        
wb.save(excel_file_path)


# In[31]:


# Split values in column B using "/" and place each part in separate columns
max_split_count = 0
for row in ws.iter_rows(min_col=2, max_col=2, min_row=2, max_row=len(lines)):
    cell = row[0]
    if cell.value:
        parts = cell.value.split("/")
        max_split_count = max(max_split_count, len(parts))
        for col_offset, part in enumerate(parts, start=1):
            ws.cell(row=cell.row, column=2 + col_offset, value=part)
wb.save(excel_file_path)


# In[32]:


# Delete columns B and C
ws.delete_cols(2, 2)

# Remove "delivery note " from values in column D
for row_num in range(2, len(lines) + 1):
    cell_value = ws.cell(row=row_num, column=2).value
    if cell_value and "delivery note " in cell_value:
        ws.cell(row=row_num, column=2).value = cell_value.replace("delivery note ", "")
wb.save(excel_file_path)


# In[33]:


# Delete rows with empty cells in column A
rows_to_delete = []
for row_num in range(1, len(lines) + 1):
    cell_value_A = ws.cell(row=row_num, column=1).value
    if not cell_value_A:
        rows_to_delete.append(row_num)

for row_num in reversed(rows_to_delete):
    ws.delete_rows(row_num)


# In[34]:


# Find formats in column A and copy the cell below to column E
for row_num in range(2, len(lines) + 1):
    cell_value_A = ws.cell(row=row_num, column=1).value
    if cell_value_A:
        for pattern, explanation in formats:
            adjusted_pattern = pattern.replace("A", "[A-Za-z]")
            if re.search(adjusted_pattern, cell_value_A):
                ws.cell(row=row_num, column=5).value = ws.cell(row=row_num + 1, column=1).value
                break


# In[35]:


# Fill empty cells in columns B, C, and D with cells above them
for col_num in range(2, 5):  # Columns B, C, and D
    current_value = None
    for row_num in range(2, len(lines) + 1):
        cell_value_col = ws.cell(row=row_num, column=col_num).value
        if cell_value_col:
            current_value = cell_value_col
        elif current_value:
            ws.cell(row=row_num, column=col_num).value = current_value
            


# In[36]:


# Copy cells from column A to the first cell in column F if they start with "Cust. no.:" and have more than 40 characters
for row_num in range(2, len(lines) + 1):
    cell_value_A = ws.cell(row=row_num, column=1).value
    if cell_value_A and cell_value_A.startswith("Cust. no.:") and len(cell_value_A) > 40:
        ws.cell(row=2, column=6).value = cell_value_A
        break


# In[37]:


# Extract characters from the 12th to the 18th position and store in the first cell of column F
for row_num in range(2, len(lines) + 1):
    cell_value_A = ws.cell(row=row_num, column=1).value
    if cell_value_A and cell_value_A.startswith("Cust. no.:") and len(cell_value_A) > 40:
        extracted_chars = cell_value_A[24:32]  # Extract characters from the 12th to 18th position
        ws.cell(row=2, column=6).value = extracted_chars
        break
        


# In[38]:


previous_value = None
for row_num in range(2, len(lines) + 1):
    cell_value_F = ws.cell(row=row_num, column=6).value
    if cell_value_F:
        previous_value = cell_value_F
    elif previous_value:
        ws.cell(row=row_num, column=6).value = previous_value
        


# In[39]:


# Delete entire rows where the specified formats are not found in column A
# Delete rows where column A is empty
rows_to_delete = []
for row_num in range(1, len(lines) + 1):
    cell_value_A = ws.cell(row=row_num, column=1).value
    if not cell_value_A or all(not re.search(pattern.replace("A", "[A-Za-z]"), cell_value_A) for pattern, _ in formats):
        rows_to_delete.append(row_num)

for row_num in reversed(rows_to_delete):
    ws.delete_rows(row_num)
    
wb.save(excel_file_path)


# In[40]:


# Add 3 columns between columns A and B
ws.insert_cols(idx=2, amount=3)

# Find formats in column D and copy the cell below to column G
for row_num in range(2, len(lines) + 1):
    cell_value_D = ws.cell(row=row_num, column=4).value
    if cell_value_D:
        for pattern, explanation in formats:
            adjusted_pattern = pattern.replace("A", "[A-Za-z]")
            if re.search(adjusted_pattern, cell_value_D):
                ws.cell(row=row_num, column=7).value = ws.cell(row=row_num + 1, column=4).value
                break


# In[41]:


# Remove multiple spaces and convert to single space in columns A and H
for row_num in range(1, len(lines) + 1):
    cell_value_A = ws.cell(row=row_num, column=1).value
    cell_value_H = ws.cell(row=row_num, column=8).value
    
    if cell_value_A:
        ws.cell(row=row_num, column=1).value = re.sub(r'\s+', ' ', cell_value_A).strip()
    
    if cell_value_H:
        ws.cell(row=row_num, column=8).value = re.sub(r'\s+', ' ', cell_value_H).strip()

wb.save(excel_file_path)


# In[42]:


# Extract value between last space and end of cell in column A and put it in column D
for row_num in range(1, len(lines) + 1):
    cell_value_A = ws.cell(row=row_num, column=1).value
    if cell_value_A:
        last_space_index = cell_value_A.rfind(" ")
        if last_space_index != -1:
            extracted_value = cell_value_A[last_space_index + 1:]
            ws.cell(row=row_num, column=4).value = extracted_value
            
wb.save(excel_file_path)


# In[43]:


# Extract value between third last and second last spaces in column A and put it in column C
for row_num in range(1, len(lines) + 1):
    cell_value_A = ws.cell(row=row_num, column=1).value
    if cell_value_A:
        space_indices = [index for index, char in enumerate(cell_value_A) if char == " "]
        if len(space_indices) >= 3:
            third_last_space_index = space_indices[-3]
            second_last_space_index = space_indices[-2]
            extracted_value = cell_value_A[third_last_space_index + 1:second_last_space_index]
            ws.cell(row=row_num, column=3).value = extracted_value

wb.save(excel_file_path)


# In[44]:


# Extract value between the fifth character and the first space in column A and put it in column B
for row_num in range(1, len(lines) + 1):
    cell_value_A = ws.cell(row=row_num, column=1).value
    if cell_value_A and len(cell_value_A) >= 5:
        first_space_index = cell_value_A.find(" ")
        if first_space_index != -1:
            extracted_value = cell_value_A[4:first_space_index]
            ws.cell(row=row_num, column=2).value = extracted_value
            
wb.save(excel_file_path)


# In[45]:


# Delete rows where column B value is empty
rows_to_delete = []
for row_num in range(1, len(lines) + 1):
    cell_value_B = ws.cell(row=row_num, column=2).value
    if not cell_value_B:
        rows_to_delete.append(row_num)

for row_num in reversed(rows_to_delete):
    ws.delete_rows(row_num)

wb.save(excel_file_path)


# In[46]:


# Delete column A
ws.delete_cols(1)
wb.save(excel_file_path)


# In[47]:


header_row = ["Part Number", "Quantity", "Unit Price", "Delivery Note P1", "Delivery Note P2", "Order Number", "Description", "Invoice No."]

# Insert the header row at the top
ws.insert_rows(1)
for col_num, value in enumerate(header_row, start=1):
    ws.cell(row=1, column=col_num, value=value)
    
wb.save(excel_file_path)


# In[48]:


# Change the column orders
column_order = ['A', 'B', 'F', 'G', 'H', 'C', 'D', 'E']
new_ws = ws.parent.create_sheet("NewSheet")
for col_idx, col_letter in enumerate(column_order, start=1):
    new_col = new_ws.column_dimensions[get_column_letter(col_idx)]
    old_col = ws.column_dimensions[col_letter]
    new_col.width = old_col.width
    new_col.hidden = old_col.hidden
    for cell in ws[col_letter]:
        new_cell = new_ws.cell(row=cell.row, column=col_idx, value=cell.value)


# In[49]:


# Rename "NewSheet" with the first value below the header from column "H"
header_value = new_ws.cell(row=2, column=5).value
ws.parent[new_ws.title].title = header_value

# Delete the "Sheet" sheet
del ws.parent['Sheet']


# In[50]:


wb.save(excel_file_path)


try:
    # Save the Excel file in the same path as the PDF
    workbook.save(excel_file_path)
    print(f"Excel file '{excel_file_name}' created with PDF text.")
except Exception as e:
    print(f"An error occurred while saving the Excel file: {e}")


# In[32]:


print(f"PDF file path: {OLD_PDF_PATH}")
print(f"Excel file path: {excel_file_path}")
