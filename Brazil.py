#!/usr/bin/env python
# coding: utf-8

# In[1]:


import PyPDF2
import openpyxl
import os
import re
from openpyxl.utils import get_column_letter
import sys


# In[2]:

OLD_PDF_PATH = r"NEW_PDF_PATH"
OLD_PDF_NAME = os.path.basename(OLD_PDF_PATH)
OLD_PDF_DIR = os.path.dirname(OLD_PDF_PATH)

# In[3]:


with open(OLD_PDF_PATH, 'rb') as pdf_file:
    pdf_reader = PyPDF2.PdfReader(pdf_file)
    
    pdf_text = ''
    for page_num in range(len(pdf_reader.pages)):
        page = pdf_reader.pages[page_num]
        pdf_text += page.extract_text()


# In[4]:


# Split the extracted text into lines
text_lines = pdf_text.split('\n')


# In[5]:

exe_dir = getattr(sys, '_MEIPASS', os.path.abspath(os.path.dirname(__file__)))
# Create an Excel workbook
excel_file_name = os.path.splitext(os.path.basename(OLD_PDF_PATH))[0] + '.xlsx'
excel_file_path = os.path.join(OLD_PDF_DIR, excel_file_name)
# Create an Excel workbook
# excel_file_name = os.path.splitext(os.path.basename(OLD_PDF_PATH))[0] + '.xlsx'
# excel_file_path = os.path.join(os.path.dirname(OLD_PDF_PATH), excel_file_name)

workbook = openpyxl.Workbook()
sheet = workbook.active
sheet.title = "hope"  # Set the title of the first sheet to "hope"


# In[6]:


# Assign each line to a separate row in the "hope" sheet
for row_num, line in enumerate(text_lines, start=1):
    sheet.cell(row=row_num, column=1, value=line)


# In[7]:


sheet = workbook.active  # Assuming you want to work with the active sheet

# List of formats to search for in Column A
formats = [
    r'[A-Za-z] \d{3} \d{3} \d{2} \d{2}',
    r'[A-Za-z] \d{3} \d{3} \d{2} \d{2} \d{2}',
    r'[A-Za-z] \d{6} \d{6}',
    r'[A-Za-z] \d{6} \d{6} \d{2}',
    r'[A-Za-z] \d{3} \d{3} \d{2} \d{2} \d{4}',
    r'[A-Za-z] \d{3} \d{3} \d{2} \d{2} \d{2} \d{4}',
    r'[A-Za-z] \d{9}',
    r'[A-Za-z] \d{9} \d{2}',
    r'[A-Za-z] \d{3} \d{3} \d{2} \d{2} [A-Za-z]\d{2}',
    r'[A-Za-z] \d{3} \d{3} \d{2} \d{2} [A-Za-z]\d{2}',
    r'[A-Za-z] \d{6} \d{6} \d{4}',
    r'[A-Za-z] \d{4} \d{4}',
    r'[A-Za-z] \d{4} \d{4} \d{2}',
    # Add more formats as needed
]

# Iterate through rows to find formats and copy data
for i in range(1, sheet.max_row + 1):
    cell_value = sheet.cell(row=i, column=1).value
    
    for format_regex in formats:
        match = re.search(format_regex, cell_value)
        if match:
            format_text = match.group()
            
            # Copy values from next rows
            for offset, col_letter in enumerate(["B", "C", "D", "E", "F", "G"], start=1):
                next_row_value = sheet.cell(row=i + offset, column=1).value
                sheet[f"{col_letter}{i}"].value = next_row_value
            break  # Exit the loop once a format is found


# In[8]:


sheet = workbook.active  # Assuming you want to work with the active sheet

# List of formats to search for in Column A
formats = [
    r'[A-Za-z] \d{3} \d{3} \d{2} \d{2}',
    r'[A-Za-z] \d{3} \d{3} \d{2} \d{2} \d{2}',
    r'[A-Za-z] \d{6} \d{6}',
    r'[A-Za-z] \d{6} \d{6} \d{2}',
    r'[A-Za-z] \d{3} \d{3} \d{2} \d{2} \d{4}',
    r'[A-Za-z] \d{3} \d{3} \d{2} \d{2} \d{2} \d{4}',
    r'[A-Za-z] \d{9}',
    r'[A-Za-z] \d{9} \d{2}',
    r'[A-Za-z] \d{3} \d{3} \d{2} \d{2} [A-Za-z]\d{2}',
    r'[A-Za-z] \d{3} \d{3} \d{2} \d{2} [A-Za-z]\d{2}',
    r'[A-Za-z] \d{6} \d{6} \d{4}',
    r'[A-Za-z] \d{4} \d{4}',
    r'[A-Za-z] \d{4} \d{4} \d{2}',
    # Add more formats as needed
]

# Iterate through rows in reverse and delete if format is not found
for i in range(sheet.max_row, 0, -1):
    cell_value = sheet.cell(row=i, column=1).value
    found_format = False
    
    for format_regex in formats:
        match = re.search(format_regex, cell_value)
        if match:
            found_format = True
            break
    
    if not found_format:
        sheet.delete_rows(i)


# In[9]:


sheet = workbook.active  # Assuming you want to work with the active sheet

# Columns A to G
columns_to_process = ['A', 'B', 'C', 'D', 'E', 'F', 'G']

# Iterate through specified columns and replace multi spaces
for col in columns_to_process:
    col_index = openpyxl.utils.column_index_from_string(col)
    for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=col_index, max_col=col_index):
        cell_value = row[0].value
        if cell_value:
            cell_value = re.sub(r'\s+', ' ', cell_value).strip()  # Replace consecutive spaces with a single space
            row[0].value = cell_value


# In[10]:


sheet = workbook.active  # Assuming you want to work with the active sheet

# Delete column G
col_to_delete = 'G'
col_index = openpyxl.utils.column_index_from_string(col_to_delete)
sheet.delete_cols(col_index)


# In[11]:


sheet = workbook.active  # Assuming you want to work with the active sheet

# Insert 5 new columns between columns A and B
num_cols_to_insert = 5
col_index = openpyxl.utils.column_index_from_string('B')  # Column B's index
sheet.insert_cols(col_index, amount=num_cols_to_insert)


# In[12]:


sheet = workbook.active  # Assuming you want to work with the active sheet

# Iterate through rows in column A and update values in column E
for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=1, max_col=1):
    cell_value = row[0].value
    last_space_index = cell_value.rfind(" ")
    second_last_space_index = cell_value.rfind(" ", 0, last_space_index - 1)
    
    if last_space_index != -1 and second_last_space_index != -1:
        value_between_spaces = cell_value[second_last_space_index + 1:last_space_index]
        sheet.cell(row=row[0].row, column=5).value = value_between_spaces


# In[13]:


sheet = workbook.active  # Assuming you want to work with the active sheet

# Iterate through rows in column A and update values in column D
for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=1, max_col=1):
    cell_value = row[0].value
    second_last_space_index = cell_value.rfind(" ", 0, cell_value.rfind(" ") - 1)
    third_last_space_index = cell_value.rfind(" ", 0, second_last_space_index - 1)
    
    if second_last_space_index != -1 and third_last_space_index != -1:
        value_between_spaces = cell_value[third_last_space_index + 1:second_last_space_index]
        sheet.cell(row=row[0].row, column=4).value = value_between_spaces


# In[14]:


sheet = workbook.active  # Assuming you want to work with the active sheet

# Iterate through rows in column A and update values in column C
for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=1, max_col=1):
    cell_value = row[0].value
    fourth_last_space_index = cell_value.rfind(" ", 0, cell_value.rfind(" ", 0, cell_value.rfind(" ") - 1) - 1)
    third_last_space_index = cell_value.rfind(" ", 0, fourth_last_space_index - 1)
    
    if fourth_last_space_index != -1 and third_last_space_index != -1:
        value_between_spaces = cell_value[third_last_space_index + 1:fourth_last_space_index]
        sheet.cell(row=row[0].row, column=3).value = value_between_spaces


# In[15]:


sheet = workbook.active  # Assuming you want to work with the active sheet

# Iterate through rows in column A and update values in column B
for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=1, max_col=1):
    cell_value = row[0].value
    fourth_last_space_index = cell_value.rfind(" ", 0, cell_value.rfind(" ", 0, cell_value.rfind(" ", 0, cell_value.rfind(" ") - 1) - 1) - 1)
    first_space_index = cell_value.find(" ")
    
    if first_space_index != -1 and fourth_last_space_index != -1:
        value_between_spaces = cell_value[first_space_index + 1:fourth_last_space_index]
        sheet.cell(row=row[0].row, column=2).value = value_between_spaces


# In[16]:


sheet = workbook.active  # Assuming you want to work with the active sheet

# Iterate through rows in column A and update values in column F
for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=1, max_col=1):
    cell_value = row[0].value
    last_space_index = cell_value.rfind(" ")
    
    if last_space_index != -1:
        value_after_last_space = cell_value[last_space_index + 1:]
        sheet.cell(row=row[0].row, column=6).value = value_after_last_space


# In[17]:


sheet = workbook.active  # Assuming you want to work with the active sheet

# Insert a new column between columns H and I
insert_col_index = 9  # Index where you want to insert the new column
sheet.insert_cols(insert_col_index)


# In[18]:


sheet = workbook.active  # Assuming you want to work with the active sheet

# Iterate through cells in column H and remove substrings
for row in sheet.iter_rows(min_row=1, min_col=8, max_col=8):
    cell = row[0]
    if cell.value and isinstance(cell.value, str):  # Check if the cell contains a string value
        modified_value = cell.value.replace("NCM/SH: ", "").replace("NALADI/SH: ", "")
        cell.value = modified_value


# In[19]:


sheet = workbook.active  # Assuming you want to work with the active sheet

# Iterate through cells in column H and split values
for row in sheet.iter_rows(min_row=1, min_col=8, max_col=8):
    cell = row[0]
    if cell.value and isinstance(cell.value, str):  # Check if the cell contains a string value
        parts = cell.value.split(" ")  # Split the value by spaces
        if len(parts) >= 2:
            cell.value = parts[0]  # Assign the first part to column H
            sheet.cell(row=cell.row, column=9).value = " ".join(parts[1:])  # Assign the rest to column I


# In[20]:


sheet = workbook.active  # Assuming you want to work with the active sheet

# Iterate through cells in column J and remove "ORDER/PEDIDO "
for row in sheet.iter_rows(min_row=1, min_col=10, max_col=10):
    cell = row[0]
    if cell.value and isinstance(cell.value, str):  # Check if the cell contains a string value
        cell.value = cell.value.replace("ORDER/PEDIDO ", "")

# Iterate through cells in column K and remove "ORIGIN/ORIGEN: "
for row in sheet.iter_rows(min_row=1, min_col=11, max_col=11):
    cell = row[0]
    if cell.value and isinstance(cell.value, str):  # Check if the cell contains a string value
        cell.value = cell.value.replace("ORIGIN/ORIGEN: ", "")


# In[21]:


# Iterate through cells in column L and split by "("
for row in sheet.iter_rows(min_row=1, min_col=12, max_col=12):
    cell = row[0]
    if cell.value and isinstance(cell.value, str):  # Check if the cell contains a string value
        parts = cell.value.split("(")
        if len(parts) >= 2:
            cell.value = parts[0].strip()  # Set the value in column L
            sheet.cell(row=cell.row, column=13, value=parts[1].strip())  # Set the value in column M


# In[22]:


sheet = workbook.active  # Assuming you want to work with the active sheet

# Iterate through cells in column M and remove ")"
for row in sheet.iter_rows(min_row=1, min_col=13, max_col=13):
    cell = row[0]
    if cell.value and isinstance(cell.value, str):  # Check if the cell contains a string value
        cell.value = cell.value.replace(")", "").strip()  # Remove ")" and strip any whitespace


# In[23]:


sheet = workbook.active  # Assuming you want to work with the active sheet

# Delete column A
sheet.delete_cols(1)  # Delete the first column


# In[24]:


sheet = workbook.active  # Assuming you want to work with the active sheet

# Column names for the new header
new_header = [
    "Part Number",
    "Net Weight",
    "Quantity",
    "Unit Price",
    "Total Amount",
    "Description",
    "NCM",
    "NALADI",
    "Order Number",
    "Origin",
    "Code",
    "Alternative Part Number"
]

# Insert the new header above the first row
sheet.insert_rows(1)

# Populate the new header with column names
for col_index, col_name in enumerate(new_header, start=1):
    sheet.cell(row=1, column=col_index).value = col_name


# In[25]:


sheet = workbook.active  # Assuming you want to work with the active sheet

# Specify the new order of column indices
new_column_order = [1, 12, 3, 6, 4, 5, 7, 8, 9, 10, 11, 2]

# Create a new worksheet to store the reordered data
reordered_sheet = workbook.create_sheet("Reordered")

# Iterate through rows in the original sheet and copy data to the reordered sheet
for row_index, row in enumerate(sheet.iter_rows(), start=1):
    for col_index, new_col in enumerate(new_column_order, start=1):
        reordered_sheet.cell(row=row_index, column=col_index, value=row[new_col - 1].value)

# Remove the original sheet
workbook.remove(sheet)


# In[26]:


try:
    # Save the Excel file in the same path as the PDF
    workbook.save(excel_file_path)
    print(f"Excel file '{excel_file_name}' created with PDF text.")
except Exception as e:
    print(f"An error occurred while saving the Excel file: {e}")


# In[27]:


print(f"PDF file path: {OLD_PDF_PATH}")
print(f"Excel file path: {excel_file_path}")


# In[ ]:




